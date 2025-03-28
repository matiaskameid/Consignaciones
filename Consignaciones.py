import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import zipfile
import re
import math
import json

st.title("Liquidaciones de Consignaciones")

# Archivo JSON para persistir los datos de contacto
CONTACT_DATA_FILE = "contact_data.json"

def load_contact_data():
    """Intenta leer el archivo JSON y retorna un diccionario. Si no existe, retorna {}."""
    if os.path.exists(CONTACT_DATA_FILE):
        try:
            with open(CONTACT_DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data
        except Exception as e:
            st.error(f"Error al cargar {CONTACT_DATA_FILE}: {e}")
            return {}
    else:
        return {}

def save_contact_data(data):
    """Guarda el diccionario 'data' en el archivo JSON."""
    try:
        with open(CONTACT_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        st.success("Datos de contacto guardados correctamente.")
    except Exception as e:
        st.error(f"Error al guardar {CONTACT_DATA_FILE}: {e}")

# Función para cargar el logo por defecto
def load_default_logo():
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(current_dir, "logo.png")
        if os.path.exists(logo_path):
            with open(logo_path, "rb") as f:
                return BytesIO(f.read())
    except Exception:
        pass
    return None

def create_export_excel(df, editorial, logo_content=None, contact_info=None):
    """
    Genera el archivo Excel para la editorial dada, con el formato solicitado.
    Parámetros:
      - df: DataFrame con columnas [Unidades a liquidar, Producto, ISBN].
      - editorial: nombre de la editorial (en mayúsculas).
      - logo_content: bytes del logo (si existe).
      - contact_info: diccionario con datos de contacto: 
            "PROVEEDOR", "CONTACTO", "FONO / MAIL", "DESCUENTO", "PAGO", "FECHA"
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"
    ws.sheet_view.showGridLines = False

    # Estilos
    title_font = Font(name="Arial", size=16, bold=True)
    header_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Altura de la primera fila para el logo
    ws.row_dimensions[1].height = 45

    # Insertar logo (si existe)
    if logo_content is not None:
        try:
            logo_bytes = BytesIO(logo_content)
            img = OpenpyxlImage(logo_bytes)
            img.width = 80
            img.height = 50
            ws.add_image(img, "A1")
        except Exception:
            pass

    # Título en celdas fusionadas B1:D2
    ws.merge_cells("B1:D2")
    cell_title = ws["B1"]
    title_text = f"LIQUIDACION CONSIGNACIONES {editorial}"
    cell_title.value = title_text
    cell_title.font = title_font
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    # Información de cliente (RUT actualizado)
    ws.merge_cells("B3:D6")
    cell_cliente = ws["B3"]
    cell_cliente.value = (
        "CLIENTE: Librería Virtual y Distribuidora El Ático Ltda.\n"
        "Venta y Distribución de Libros\n"
        "General Bari 234, Providencia - Santiago, Teléfono: (56)2 21452308\n"
        "Rut: 76082908-0"
    )
    cell_cliente.font = normal_font
    cell_cliente.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # Tabla de proveedor/contacto (filas 8 a 13)
    fields = ["PROVEEDOR:", "CONTACTO:", "FONO / MAIL:", "DESCUENTO:", "PAGO:", "FECHA:"]
    row_start_contact = 8
    for i, field in enumerate(fields):
        row_i = row_start_contact + i
        ws.cell(row=row_i, column=2, value=field).font = header_font
        ws.merge_cells(start_row=row_i, start_column=3, end_row=row_i, end_column=4)
        merged_cell = ws.cell(row=row_i, column=3)
        key = field.replace(":", "")
        merged_cell.value = contact_info.get(key, "") if contact_info else ""
        merged_cell.font = normal_font
        for c in range(2, 5):
            ws.cell(row=row_i, column=c).border = thin_border

    # Tabla de datos a partir de la fila 16 (columnas B, C, D)
    start_row = 16
    start_col = 2
    headers = list(df.columns)  # ["Unidades a liquidar", "Producto", "ISBN"]
    for offset, header in enumerate(headers):
        col_index = start_col + offset
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_counter = start_row
    for row_data in df.itertuples(index=False):
        row_counter += 1
        for offset, value in enumerate(row_data):
            col_index = start_col + offset
            cell = ws.cell(row=row_counter, column=col_index, value=value)
            if offset == 2:  # Columna ISBN
                try:
                    cell.value = int(value)
                except Exception:
                    cell.value = value
                cell.number_format = "0"
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

    # Ajuste de anchos de columnas en la tabla de datos:
    units_header = "Unidades a liquidar"
    units_width = len(units_header) + 2

    product_header = "Producto"
    if not df.empty:
        max_product_length = max(len(str(x)) for x in df["Producto"])
    else:
        max_product_length = len(product_header)
    product_width = max(len(product_header), max_product_length) + 5

    isbn_width = 15

    required_total_title_width = 2 * len(title_text)
    current_total_width = units_width + product_width + isbn_width
    if required_total_title_width > current_total_width:
        extra = required_total_title_width - current_total_width
        product_width += extra

    ws.column_dimensions[get_column_letter(2)].width = units_width
    ws.column_dimensions[get_column_letter(3)].width = product_width
    ws.column_dimensions[get_column_letter(4)].width = isbn_width

    max_col = ws.max_column
    if max_col > 5:
        ws.delete_cols(6, max_col - 5)
    for col in range(6, 16385):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def process_master_file(file_bytes, logo_content=None, contact_infos=None):
    """
    Procesa el archivo maestro para generar liquidaciones.
    Retorna (output_files, no_data_editorials):
      - output_files: dict de archivos generados.
      - no_data_editorials: lista de editoriales sin unidades a liquidar.
    """
    df = pd.read_excel(file_bytes, sheet_name=0, header=5)
    df.columns = df.columns.str.strip()
    
    if "Código" in df.columns:
        df.rename(columns={"Código": "Codigo"}, inplace=True)

    all_cols = df.columns.tolist()
    consign_cols = [col for col in all_cols if re.search(r'consignacion', col, re.IGNORECASE)]
    
    output_files = {}
    no_data_editorials = []

    for col in consign_cols:
        editorial_name = re.sub(r'(?i)consignacion(es)?', '', col)
        editorial_name = re.sub(r'\s+', ' ', editorial_name)
        editorial_name = re.sub(r'[:]+', '', editorial_name)
        editorial_name = re.sub(r'[0-9-]+', '', editorial_name)
        editorial_name = editorial_name.strip().upper()
        if not editorial_name:
            editorial_name = "SIN EDITORIAL"

        required_cols = ["Producto", "Codigo", "BODEGA GENERAL BARI"]
        if not all(x in df.columns for x in required_cols):
            st.error("Faltan columnas requeridas en el archivo maestro.")
            return {}, []

        temp_df = df[["Producto", "Codigo", "BODEGA GENERAL BARI", col]].copy()
        temp_df.rename(columns={col: "Consignaciones"}, inplace=True)
        temp_df = temp_df[temp_df["BODEGA GENERAL BARI"] >= 0]
        
        temp_df["Unidades a liquidar"] = temp_df["Consignaciones"] - temp_df["BODEGA GENERAL BARI"]
        temp_df = temp_df[temp_df["Unidades a liquidar"] > 0]
        temp_df = temp_df.sort_values(by="Producto")
        
        if temp_df.empty:
            no_data_editorials.append(editorial_name)
            continue
        
        export_df = temp_df[["Unidades a liquidar", "Producto", "Codigo"]].copy()
        export_df.rename(columns={"Codigo": "ISBN"}, inplace=True)
        export_df["ISBN"] = export_df["ISBN"].astype(str).apply(lambda x: x.split("/")[0][:13])
        
        contact_info = contact_infos.get(editorial_name, {}) if contact_infos else {}
        excel_bytes = create_export_excel(export_df, editorial_name, logo_content, contact_info)
        filename = f"Liquidacion_Consignaciones_{editorial_name}.xlsx"
        output_files[filename] = excel_bytes

    return output_files, no_data_editorials

# Cargar datos de contacto del JSON
contact_data = load_contact_data()

# Subir archivo maestro
uploaded_file = st.file_uploader("Sube el archivo Excel maestro (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    file_bytes = BytesIO(uploaded_file.read())
    file_bytes.seek(0)
    df_temp = pd.read_excel(file_bytes, sheet_name=0, header=5)
    df_temp.columns = df_temp.columns.str.strip()
    all_cols = df_temp.columns.tolist()
    consign_cols = [col for col in all_cols if re.search(r'consignacion', col, re.IGNORECASE)]
    editorial_names = []
    for col in consign_cols:
        ename = re.sub(r'(?i)consignacion(es)?', '', col)
        ename = re.sub(r'\s+', ' ', ename)
        ename = re.sub(r'[:]+', '', ename)
        ename = re.sub(r'[0-9-]+', '', ename)
        ename = ename.strip().upper()
        if not ename:
            ename = "SIN EDITORIAL"
        editorial_names.append(ename)
    unique_editorials = sorted(list(set(editorial_names)))
    
    st.header("Completa los datos de contacto para cada editorial")
    contact_infos = {}
    for ed in unique_editorials:
        # Usar el valor guardado en el JSON o dejar vacío
        default = contact_data.get(ed, {})
        with st.expander(f"Datos para {ed}"):
            proveedor = st.text_input(f"Proveedor para {ed}", value=default.get("PROVEEDOR", ""), key=f"proveedor_{ed}")
            contacto = st.text_input(f"Contacto para {ed}", value=default.get("CONTACTO", ""), key=f"contacto_{ed}")
            fono_mail = st.text_input(f"Fono / Mail para {ed}", value=default.get("FONO / MAIL", ""), key=f"fono_mail_{ed}")
            descuento = st.text_input(f"Descuento para {ed}", value=default.get("DESCUENTO", ""), key=f"descuento_{ed}")
            pago = st.text_input(f"Pago para {ed}", value=default.get("PAGO", ""), key=f"pago_{ed}")
            fecha = st.text_input(f"Fecha para {ed}", value=default.get("FECHA", ""), key=f"fecha_{ed}")  # Puedes usar st.date_input si prefieres
            contact_infos[ed] = {
                "PROVEEDOR": proveedor,
                "CONTACTO": contacto,
                "FONO / MAIL": fono_mail,
                "DESCUENTO": descuento,
                "PAGO": pago,
                "FECHA": fecha
            }
    
    if st.button("Guardar Contactos"):
        # Actualizar el archivo JSON con los nuevos datos
        updated_data = contact_data.copy()
        updated_data.update(contact_infos)
        save_contact_data(updated_data)
        # Actualizamos la variable para la sesión
        contact_data = updated_data

    if st.button("Generar Liquidaciones"):
        file_bytes.seek(0)
        default_logo = load_default_logo()
        logo_data = default_logo.read() if default_logo else None
        results, no_data_editorials = process_master_file(file_bytes, logo_data, contact_infos)
        
        if results:
            st.success("Liquidaciones generadas para las siguientes editoriales:")
            for name in results.keys():
                st.write("📄", name)
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for filename, content in results.items():
                    zip_file.writestr(filename, content)
            zip_buffer.seek(0)
            st.download_button(
                label="📦 Descargar ZIP con todas las liquidaciones",
                data=zip_buffer,
                file_name="Liquidaciones.zip",
                mime="application/zip"
            )
        else:
            st.error("No se generaron liquidaciones (posiblemente no hay registros con unidades a liquidar > 0).")

        if no_data_editorials:
            st.info("No se generaron liquidaciones para las siguientes editoriales porque no hay unidades a liquidar:")
            for nd in no_data_editorials:
                st.write("-", nd)
